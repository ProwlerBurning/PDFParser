import sys
from argparse import Namespace
from pathlib import Path

import fitz
from openpyxl import load_workbook

import extract
from src.export_excel import export_workbook
from src.parsers.base import ParseResult, transaction_row
from src.pdf_detect import detect_statement_type, processing_mode_for
from src.parser_builder import ParserBuildError


def test_detects_all_supported_statement_types():
    assert detect_statement_type("TNG WALLET TRANSACTION HISTORY") == "tng"
    assert detect_statement_type(
        "transaction date time transaction type transaction amount "
        "transaction direction transaction status transaction reference id"
    ) == "tng"
    assert detect_statement_type("Standard Chartered Credit Card Statement") == "sc"
    assert detect_statement_type("YOUR TRANSACTION DETAILS ESSENTIAL VISA CLASSIC") == "hlb"


def test_hlb_uses_text_and_sc_tng_use_ocr():
    assert processing_mode_for("hlb", has_useful_text=True) == "text"
    assert processing_mode_for("tng", has_useful_text=True) == "text"
    assert processing_mode_for("tng", has_useful_text=False) == "ocr"
    assert processing_mode_for("sc", has_useful_text=False) == "ocr"


def test_export_has_required_sheets(tmp_path: Path):
    output = tmp_path / "result.xlsx"
    export_workbook(
        output,
        transactions=[],
        summaries=[],
        exceptions=[],
        raw_extract=[],
    )
    workbook = load_workbook(output, read_only=True)
    assert workbook.sheetnames == [
        "Transactions",
        "Statement_Summary",
        "Exceptions",
        "Raw_Extract",
    ]


def test_tng_workbook_privacy_respects_unmask_flag(tmp_path: Path, monkeypatch):
    pdf_path = tmp_path / "tng.pdf"
    pdf_path.write_bytes(b"placeholder")
    source_text = (
        "TNG WALLET TRANSACTION HISTORY\n"
        "Registered Name: Example Person\n"
        "Wallet ID: 60123456789\n"
        "1/6/2025 Success Payment 123456789012 Sample Merchant "
        "RM10.00 RM90.00"
    )
    monkeypatch.setattr(extract, "extract_native_pages", lambda *args: ([], False))
    monkeypatch.setattr(
        extract,
        "ocr_pdf_pages",
        lambda *args, **kwargs: ([{"text": source_text}], False),
    )
    logger = RecordingLogger()

    masked = extract.process_pdf(pdf_path, tmp_path / "cache", None, False, False, logger)
    unmasked = extract.process_pdf(pdf_path, tmp_path / "cache", None, False, True, logger)
    masked_path = tmp_path / "masked.xlsx"
    unmasked_path = tmp_path / "unmasked.xlsx"
    for path, result in ((masked_path, masked), (unmasked_path, unmasked)):
        export_workbook(
            path,
            result.transactions,
            result.summaries,
            result.exceptions,
            result.raw_extract,
        )

    masked_book = load_workbook(masked_path, read_only=True, data_only=True)
    unmasked_book = load_workbook(unmasked_path, read_only=True, data_only=True)
    masked_headers = [cell.value for cell in next(masked_book["Transactions"].iter_rows())]
    masked_values = dict(
        zip(masked_headers, next(masked_book["Transactions"].iter_rows(min_row=2, values_only=True)))
    )
    unmasked_headers = [cell.value for cell in next(unmasked_book["Transactions"].iter_rows())]
    unmasked_values = dict(
        zip(unmasked_headers, next(unmasked_book["Transactions"].iter_rows(min_row=2, values_only=True)))
    )
    masked_summary_headers = [
        cell.value for cell in next(masked_book["Statement_Summary"].iter_rows())
    ]
    masked_summary = dict(
        zip(
            masked_summary_headers,
            next(masked_book["Statement_Summary"].iter_rows(min_row=2, values_only=True)),
        )
    )
    unmasked_summary_headers = [
        cell.value for cell in next(unmasked_book["Statement_Summary"].iter_rows())
    ]
    unmasked_summary = dict(
        zip(
            unmasked_summary_headers,
            next(unmasked_book["Statement_Summary"].iter_rows(min_row=2, values_only=True)),
        )
    )

    assert masked_values["account_name_masked"] == "E***** P*****"
    assert masked_values["account_no_masked"] == "601*****789"
    assert masked_values["reference"] != "123456789012"
    assert masked_values["raw_line"] == "[MASKED RAW LINE - use --unmask for local review]"
    assert unmasked_values["account_name_masked"] == "Example Person"
    assert unmasked_values["account_no_masked"] == "60123456789"
    assert unmasked_values["reference"] == "123456789012"
    assert "123456789012" in unmasked_values["raw_line"]
    assert masked_summary["account_name_masked"] == "E***** P*****"
    assert masked_summary["account_no_masked"] == "601*****789"
    assert unmasked_summary["account_name_masked"] == "Example Person"
    assert unmasked_summary["account_no_masked"] == "60123456789"


def test_parser_builder_dry_run_cli_creates_artifacts_without_workbook(
    tmp_path: Path,
    monkeypatch,
):
    pdf_path = tmp_path / "unknown_123456789012.pdf"
    document = fitz.open()
    page = document.new_page()
    page.insert_text(
        (72, 72),
        "Example Community Bank\nAccount Number: 123456789012\n"
        "01 JAN SAMPLE SHOP 10.00",
    )
    document.save(pdf_path)
    document.close()
    artifact_dir = tmp_path / "parser_build"
    workbook_path = tmp_path / "should_not_exist.xlsx"
    monkeypatch.delenv("OPENROUTER_API_KEY", raising=False)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "extract.py",
            "--input",
            str(pdf_path),
            "--output",
            str(workbook_path),
            "--parser-build-dry-run",
            "--parser-build-output",
            str(artifact_dir),
        ],
    )

    assert extract.main() == 0
    prompt_files = list(artifact_dir.glob("*.prompt.txt"))
    assert prompt_files
    assert list(artifact_dir.glob("*.metadata.json"))
    prompt = prompt_files[0].read_text()
    assert "123456789012" not in prompt
    assert not workbook_path.exists()


def test_normal_extraction_does_not_enter_parser_builder(tmp_path: Path, monkeypatch):
    pdf_path = tmp_path / "known.pdf"
    pdf_path.write_bytes(b"placeholder")
    output_path = tmp_path / "normal.xlsx"
    result = ParseResult(
        transactions=[
            transaction_row(
                source_file="known.pdf",
                provider="Hong Leong Bank",
                amount=-1.0,
            )
        ],
        summaries=[
            {
                "source_file": "known.pdf",
                "provider": "Hong Leong Bank",
                "processing_mode": "text",
                "transaction_count": 1,
            }
        ],
    )

    monkeypatch.setattr(extract, "collect_pdfs", lambda path: [pdf_path])
    monkeypatch.setattr(extract, "process_pdf", lambda *args, **kwargs: result)
    monkeypatch.setattr(
        extract,
        "run_parser_build_mode",
        lambda *args, **kwargs: (_ for _ in ()).throw(
            AssertionError("Parser Builder must not run")
        ),
        raising=False,
    )
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "extract.py",
            "--input",
            str(pdf_path),
            "--output",
            str(output_path),
        ],
    )

    assert extract.main() == 0
    assert output_path.is_file()


class RecordingLogger:
    def __init__(self):
        self.errors = []

    def info(self, *args, **kwargs):
        pass

    def error(self, message, *args):
        self.errors.append(message % args)


def parser_build_args(tmp_path: Path) -> Namespace:
    return Namespace(
        input=tmp_path / "unknown.pdf",
        parser_build_output=tmp_path / "artifacts",
        parser_build_max_pages=5,
        parser_build_model="test/model",
        parser_build_dry_run=True,
        force=False,
        force_parser_build=False,
    )


def test_parser_builder_broad_error_logs_type_without_sensitive_message(
    tmp_path: Path,
    monkeypatch,
):
    logger = RecordingLogger()
    pdf_path = tmp_path / "unknown.pdf"
    sensitive_message = "/private/customer/123456789012.pdf"
    monkeypatch.setattr(
        extract,
        "prepare_parser_build_source",
        lambda *args, **kwargs: (_ for _ in ()).throw(
            OSError(sensitive_message)
        ),
    )

    result = extract.run_parser_build_mode(
        parser_build_args(tmp_path),
        [pdf_path],
        tmp_path,
        logger,
    )

    assert result == 2
    assert logger.errors == ["Parser build failed for unknown.pdf: OSError"]
    assert sensitive_message not in logger.errors[0]


def test_parser_builder_controlled_error_logs_safe_message(tmp_path: Path, monkeypatch):
    logger = RecordingLogger()
    pdf_path = tmp_path / "unknown.pdf"
    monkeypatch.setattr(
        extract,
        "prepare_parser_build_source",
        lambda *args, **kwargs: (_ for _ in ()).throw(
            ParserBuildError("supported statement requires --force-parser-build")
        ),
    )

    result = extract.run_parser_build_mode(
        parser_build_args(tmp_path),
        [pdf_path],
        tmp_path,
        logger,
    )

    assert result == 2
    assert logger.errors == [
        "Parser build failed for unknown.pdf: "
        "supported statement requires --force-parser-build"
    ]
