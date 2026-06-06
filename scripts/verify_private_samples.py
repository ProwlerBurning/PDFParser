#!/usr/bin/env python3
"""Locally verify extraction + privacy masking for PDF statements you supply.

Pass one or more PDF paths as positional arguments. Nothing is written
permanently, nothing is uploaded, and no online/LLM service is contacted: each
PDF is processed by the normal local deterministic pipeline, exported to a
temporary workbook, checked for the required sheets, and scanned for leaked
long digit runs. Provider/exception counts, processing modes, and the privacy
scan result are printed.

Usage:
    python scripts/verify_private_samples.py STATEMENT.pdf [MORE.pdf ...]
"""
from __future__ import annotations

import argparse
import re
import sys
import tempfile
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from extract import process_pdf
from src.export_excel import export_workbook


REQUIRED_SHEETS = [
    "Transactions",
    "Statement_Summary",
    "Exceptions",
    "Raw_Extract",
]


class SilentLogger:
    def info(self, *args, **kwargs) -> None:
        pass


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Locally verify extraction and privacy masking for the PDF "
        "statements you pass in. No files are written permanently and no "
        "online/LLM service is contacted.",
    )
    parser.add_argument(
        "pdfs",
        nargs="*",
        type=Path,
        metavar="PDF",
        help="One or more PDF statement paths to verify locally.",
    )
    args = parser.parse_args(argv)
    if not args.pdfs:
        parser.print_usage(sys.stderr)
        parser.exit(2, "error: provide at least one PDF path to verify\n")
    return args


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    missing = [path for path in args.pdfs if not path.is_file()]
    if missing:
        for path in missing:
            print(f"error: not a file: {path}", file=sys.stderr)
        return 2

    transactions = []
    summaries = []
    exceptions = []
    raw_extract = []
    logger = SilentLogger()

    with tempfile.TemporaryDirectory(prefix="statement-parser-verify-") as directory:
        # Keep the cache and the workbook inside the temporary directory so this
        # helper never writes permanent cache or output artifacts.
        temp_root = Path(directory)
        cache_root = temp_root / "cache"
        workbook_path = temp_root / "statements.xlsx"
        for path in args.pdfs:
            result = process_pdf(path, cache_root, None, False, False, logger)
            transactions.extend(result.transactions)
            summaries.extend(result.summaries)
            exceptions.extend(result.exceptions)
            raw_extract.extend(result.raw_extract)

        export_workbook(
            workbook_path,
            transactions,
            summaries,
            exceptions,
            raw_extract,
        )
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        if workbook.sheetnames != REQUIRED_SHEETS:
            print(
                f"error: unexpected sheets {workbook.sheetnames}, "
                f"expected {REQUIRED_SHEETS}",
                file=sys.stderr,
            )
            return 3
        privacy_passed = not any(
            isinstance(cell.value, str) and re.search(r"\d{9,}", cell.value)
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
        )

    transaction_counts = Counter(row["provider"] for row in transactions)
    exception_counts = Counter(row["provider"] for row in exceptions)
    modes = {row["provider"]: row["processing_mode"] for row in summaries}
    for provider in sorted(transaction_counts):
        print(
            f"provider={provider} "
            f"transactions={transaction_counts[provider]} "
            f"exceptions={exception_counts[provider]} "
            f"mode={modes.get(provider)}"
        )
    print(f"privacy_scan={'pass' if privacy_passed else 'fail'}")
    return 0 if privacy_passed else 4


if __name__ == "__main__":
    raise SystemExit(main())
